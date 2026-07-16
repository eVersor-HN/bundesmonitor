"""Parser fuer namentliche Abstimmungen.

Zwei Eingaben:
1. Die Ajax-Liste von bundestag.de (HTML-Tabelle) -> Listeneintraege mit
   Titel, Datum und XLSX/PDF-Links.
2. Die amtliche XLSX-Ergebnisdatei -> Gesamtsummen, Aggregat je Fraktion
   und Einzelstimmen aller Abgeordneten.

Es wird nichts interpretiert: keine "angenommen/abgelehnt"-Ableitung,
nur die amtlichen Zahlen.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any

from openpyxl import load_workbook

# Ajax-Endpunkt der amtlichen Abstimmungsliste (HTML-Fragment).
LIST_URL = (
    "https://www.bundestag.de/ajax/filterlist/de/parlament/plenum/abstimmung/liste/"
    "462112-462112"
)
PAGE_URL = "https://www.bundestag.de/parlament/plenum/abstimmung/liste"

_ROW_RE = re.compile(r"<tr>(.*?)</tr>", re.DOTALL)
_STRONG_RE = re.compile(r"<strong>(.*?)</strong>", re.DOTALL)
_XLSX_RE = re.compile(r'href="([^"]+\.xlsx?)"')
_PDF_RE = re.compile(r'href="([^"]+\.pdf)"')
_TAG_RE = re.compile(r"<[^>]+>")
_DATE_PREFIX_RE = re.compile(r"^(\d{2})\.(\d{2})\.(\d{4}):?\s*")
_FILE_DATE_RE = re.compile(r"/(\d{4})(\d{2})(\d{2})_")


@dataclass(frozen=True)
class VoteListEntry:
    title: str
    vote_date: date | None
    xlsx_url: str
    pdf_url: str | None


def _clean_text(html: str) -> str:
    text = _TAG_RE.sub(" ", html)
    return re.sub(r"\s+", " ", text).strip()


# Redaktionelle Tippfehler (z. B. Jahr 2062) nicht uebernehmen: Datum muss
# zwischen Gruendung des Bundestages und naher Zukunft liegen, sonst unklar.
def _plausible(value: date) -> bool:
    return date(1949, 1, 1) <= value <= date.today() + timedelta(days=45)


def parse_vote_list(html: str) -> list[VoteListEntry]:
    """Extrahiert Abstimmungs-Eintraege aus dem Ajax-HTML."""
    entries: list[VoteListEntry] = []
    for row_match in _ROW_RE.finditer(html):
        row = row_match.group(1)
        xlsx = _XLSX_RE.search(row)
        if xlsx is None:
            continue
        strong = _STRONG_RE.search(row)
        raw_title = _clean_text(strong.group(1)) if strong else ""
        vote_date: date | None = None
        date_match = _DATE_PREFIX_RE.match(raw_title)
        if date_match:
            d, m, y = date_match.groups()
            try:
                candidate = date(int(y), int(m), int(d))
            except ValueError:
                candidate = None
            if candidate is not None and _plausible(candidate):
                vote_date = candidate
            raw_title = raw_title[date_match.end() :].strip()
        if vote_date is None:
            file_match = _FILE_DATE_RE.search(xlsx.group(1))
            if file_match:
                y, m, d = file_match.groups()
                try:
                    candidate = date(int(y), int(m), int(d))
                except ValueError:
                    candidate = None
                if candidate is not None and _plausible(candidate):
                    vote_date = candidate
        pdf = _PDF_RE.search(row)
        entries.append(
            VoteListEntry(
                title=raw_title or "Namentliche Abstimmung",
                vote_date=vote_date,
                xlsx_url=xlsx.group(1),
                pdf_url=pdf.group(1) if pdf else None,
            )
        )
    return entries


@dataclass(frozen=True)
class ParsedVote:
    wahlperiode: int
    sitzung: int
    abstimm_nr: int
    totals: dict[str, int]
    by_fraktion: list[dict[str, Any]]
    einzelstimmen: list[dict[str, str]]


# Spalten der amtlichen XLSX (Kopfzeile, Kleinschreibung, Umlaut-tolerant).
_VOTE_COLS = ("ja", "nein", "enthaltung", "ungueltig", "nicht_abgegeben")


def _norm_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ü", "ue").replace("�", "ue")
    if text.startswith("ung"):
        return "ungueltig"
    if text.startswith("nichtabg"):
        return "nicht_abgegeben"
    return text


def parse_vote_xlsx(data: bytes) -> ParsedVote:
    """Parst die amtliche Ergebnis-XLSX (eine Zeile je Abgeordnetem)."""
    workbook = load_workbook(io.BytesIO(data), read_only=True)
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - openpyxl liefert immer ein Blatt
        raise ValueError("XLSX ohne Tabellenblatt")
    rows = sheet.iter_rows(values_only=True)
    header = [_norm_header(h) for h in next(rows)]
    idx = {name: header.index(name) for name in header if name}

    def col(row: tuple[Any, ...], name: str) -> Any:
        pos = idx.get(name)
        return row[pos] if pos is not None and pos < len(row) else None

    wahlperiode = sitzung = abstimm_nr = 0
    totals = dict.fromkeys(_VOTE_COLS, 0)
    fraktionen: dict[str, dict[str, int]] = {}
    einzelstimmen: list[dict[str, str]] = []

    for row in rows:
        if row is None or col(row, "bezeichnung") is None:
            continue
        wahlperiode = int(col(row, "wahlperiode") or wahlperiode)
        sitzung = int(col(row, "sitzungnr") or sitzung)
        abstimm_nr = int(col(row, "abstimmnr") or abstimm_nr)
        fraktion = str(col(row, "fraktion/gruppe") or "fraktionslos").strip()
        agg = fraktionen.setdefault(fraktion, dict.fromkeys(_VOTE_COLS, 0))
        stimme = "nicht_abgegeben"
        for name in _VOTE_COLS:
            if int(col(row, name) or 0) == 1:
                stimme = name
                break
        agg[stimme] += 1
        totals[stimme] += 1
        einzelstimmen.append(
            {
                "name": str(col(row, "bezeichnung")).strip(),
                "fraktion": fraktion,
                "stimme": stimme,
            }
        )

    by_fraktion = [
        {"fraktion": name, **counts}
        for name, counts in sorted(
            fraktionen.items(), key=lambda kv: sum(kv[1].values()), reverse=True
        )
    ]
    return ParsedVote(
        wahlperiode=wahlperiode,
        sitzung=sitzung,
        abstimm_nr=abstimm_nr,
        totals=totals,
        by_fraktion=by_fraktion,
        einzelstimmen=einzelstimmen,
    )
